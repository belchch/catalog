"""End-to-end API tests for step 06 (documents, planner WS, skills, runs).

HTTP endpoints are driven via ``TestClient``; WebSocket endpoints via
``TestClient.websocket_connect`` (see ``conftest.py`` for the approach). The
provider is a :class:`FakeProvider` whose ``script`` is populated per test.
"""

from __future__ import annotations

import asyncio
import json
from pathlib import Path

from catalog.documents.tools import build_document_tools
from catalog.llm.base import CompletionResult, ToolCall

from catalog.skills.apply import apply_skill_collect
from catalog.skills.config import SkillConfig, SkillOutput, VerifyCheck, compute_tags
from catalog.skills.repo_skill import create_skill, get_skill, update_status
from catalog.storage.repo_document import get_document
from catalog.storage.repo_message import add_message, list_messages
from catalog.storage.repo_session import get_session
from catalog.storage.repo_session_artifact import code_sha256, upsert_script_dry_run
from catalog.storage.repo_session_document import attach_documents, list_session_documents


def _completion(
    content: str | None = None, *, tool_calls: list[ToolCall] | None = None
) -> CompletionResult:
    """Build a CompletionResult with a 'stop' finish reason."""
    return CompletionResult(
        content=content,
        tool_calls=list(tool_calls or []),
        finish_reason="stop",
    )


def _build_skill_call(
    *,
    name: str = "Summarizer",
    allowed_tools: list[str] | None = None,
    verify_checks: list[dict] | None = None,
) -> ToolCall:
    """A build_skill tool call with a (by default valid) SkillConfig payload."""
    return ToolCall(
        id="build-1",
        name="build_skill",
        arguments={
            "name": name,
            "description": "Skill built from a planning session.",
            "system_prompt": "You process the document as instructed.",
            "allowed_tools": allowed_tools if allowed_tools is not None else ["read_document"],
            "model": "test/model",
            "verify_checks": verify_checks if verify_checks is not None else [{"check": "non_empty"}],
        },
    )


def _build_script_skill_call(
    *,
    code: str,
    name: str = "Uppercaser",
    verify_checks: list[dict] | None = None,
) -> ToolCall:
    """A build_skill tool call for a kind=script skill."""
    return ToolCall(
        id="build-1",
        name="build_skill",
        arguments={
            "name": name,
            "description": "A deterministic script skill.",
            "kind": "script",
            "code": code,
            "verify_checks": verify_checks if verify_checks is not None else [],
        },
    )


def _seed_committed_skill(
    db,
    *,
    name: str = "Summarizer",
    allowed_tools: list[str] | None = None,
    verify_checks: list[VerifyCheck] | None = None,
    max_retries: int = 2,
) -> str:
    """Insert a committed skill directly and return its id (for apply tests)."""
    config = SkillConfig(
        name=name,
        description="test skill",
        system_prompt="You summarize the document.",
        allowed_tools=allowed_tools if allowed_tools is not None else ["read_document"],
        model="test/model",
        max_iterations=4,
        max_retries=max_retries,
        verify_checks=verify_checks if verify_checks is not None else [],
    )
    return create_skill(
        db, name=config.name, description=config.description, config=config, status="committed"
    )


def _upload(client, filename: str, content: bytes) -> str:
    resp = client.post(
        "/documents", files={"file": (filename, content, "application/octet-stream")}
    )
    assert resp.status_code == 200, resp.text
    return resp.json()["id"]


# --------------------------------------------------------------------------- #
# Health + documents
# --------------------------------------------------------------------------- #


def test_health(client) -> None:
    resp = client.get("/health")
    assert resp.status_code == 200
    body = resp.json()
    assert body["status"] == "ok"
    assert "git_sha" in body


def test_business_endpoints_409_without_workspace(client_no_workspace) -> None:
    client = client_no_workspace
    assert client.get("/sessions").status_code == 409
    assert client.get("/documents").status_code == 409
    assert client.get("/skills").status_code == 409
    assert client.post("/export/docx", json={"doc_ids": ["x"]}).status_code == 409
    assert client.get("/settings").status_code == 200
    assert client.get("/providers").status_code == 200
    assert client.get("/models").status_code == 200


def test_close_workspace_returns_409(client, settings) -> None:
    assert client.get("/sessions").status_code == 200
    client.app.state.workspace_manager.close()
    assert client.get("/sessions").status_code == 409
    client.app.state.workspace_manager.open(Path(settings.workspace_dir), confirm_init=True)
    assert client.get("/sessions").status_code == 200


def test_upload_and_list_documents(client) -> None:
    resp = client.post(
        "/documents", files={"file": ("note.md", b"# Title\n\nbody", "text/markdown")}
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["kind"] == "md"
    assert data["title"] == "note"
    assert data["id"]
    doc_id = data["id"]

    listing = client.get("/documents")
    assert listing.status_code == 200
    ids = [d["id"] for d in listing.json()]
    assert doc_id in ids


def test_upload_unsupported_format(client) -> None:
    resp = client.post(
        "/documents", files={"file": ("bad.exe", b"MZ\x90\x00", "application/octet-stream")}
    )
    assert resp.status_code == 400
    assert "unsupported" in resp.json()["detail"].lower()


def test_upload_xls_format_hint(client, settings) -> None:
    resp = client.post(
        "/documents", files={"file": ("old.xls", b"not-xlsx", "application/octet-stream")}
    )
    assert resp.status_code == 400
    assert "пересохраните файл как .xlsx" in resp.json()["detail"]
    assert ".xls" in resp.json()["detail"]
    assert client.get("/documents").json() == []
    assert not (Path(settings.workspace_dir) / "old.xls").exists()


def test_upload_ods_format_hint(client, settings) -> None:
    resp = client.post(
        "/documents", files={"file": ("sheet.ods", b"not-xlsx", "application/octet-stream")}
    )
    assert resp.status_code == 400
    assert "пересохраните файл как .xlsx" in resp.json()["detail"]
    assert ".ods" in resp.json()["detail"]
    assert client.get("/documents").json() == []
    assert not (Path(settings.workspace_dir) / "sheet.ods").exists()


def test_upload_tsv_format_hint(client, settings) -> None:
    resp = client.post(
        "/documents", files={"file": ("table.tsv", b"a\tb\n", "text/tab-separated-values")}
    )
    assert resp.status_code == 400
    assert "пересохраните файл как .csv" in resp.json()["detail"]
    assert ".tsv" in resp.json()["detail"]
    assert client.get("/documents").json() == []
    assert not (Path(settings.workspace_dir) / "table.tsv").exists()


def test_upload_broken_xlsx(client, settings) -> None:
    resp = client.post(
        "/documents",
        files={"file": ("broken.xlsx", b"not-a-zip", "application/octet-stream")},
    )
    assert resp.status_code == 400
    assert "xlsx" in resp.json()["detail"].lower()
    listing = client.get("/documents")
    assert listing.status_code == 200
    assert listing.json() == []
    assert not (Path(settings.workspace_dir) / "broken.xlsx").exists()


def test_upload_empty_csv(client) -> None:
    resp = client.post(
        "/documents",
        files={"file": ("empty.csv", b"", "text/csv")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["kind"] == "csv"
    ids = [d["id"] for d in client.get("/documents").json()]
    assert data["id"] in ids


def test_upload_csv(client) -> None:
    resp = client.post(
        "/documents",
        files={"file": ("sample.csv", b"name,city\nAnna,Moscow\n", "text/csv")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["kind"] == "csv"
    assert data["title"] == "sample"


def test_upload_xlsx(client) -> None:
    xlsx_bytes = (Path(__file__).parent / "fixtures" / "sample.xlsx").read_bytes()
    resp = client.post(
        "/documents",
        files={
            "file": (
                "sample.xlsx",
                xlsx_bytes,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["kind"] == "xlsx"
    assert data["title"] == "sample"


def test_upload_pdf(client) -> None:
    pdf_bytes = (Path(__file__).parent / "fixtures" / "sample-text.pdf").read_bytes()
    resp = client.post(
        "/documents",
        files={"file": ("sample-text.pdf", pdf_bytes, "application/pdf")},
    )
    assert resp.status_code == 200
    data = resp.json()
    assert data["kind"] == "pdf"
    assert data["title"] == "sample-text"


_FIXTURES = Path(__file__).parent / "fixtures"


def test_extract_csv() -> None:
    from catalog.documents.extract import extract_text

    text = extract_text(str(_FIXTURES / "sample.csv"), "csv")
    assert "name,city,score" in text
    assert "Москва" in text
    assert "Казань" in text


def test_extract_csv_cp1251_fallback(tmp_path) -> None:
    from catalog.documents.extract import extract_text

    cp1251_bytes = "name,city\nAnna,Тверь\n".encode("cp1251")
    p = tmp_path / "win.csv"
    p.write_bytes(cp1251_bytes)
    text = extract_text(str(p), "csv")
    assert "Тверь" in text


def test_extract_csv_utf8_bom(tmp_path) -> None:
    from catalog.documents.extract import extract_text

    p = tmp_path / "bom.csv"
    p.write_bytes("\ufeffname,city\nAnna,Moscow\n".encode("utf-8"))
    text = extract_text(str(p), "csv")
    assert text.startswith("name,city")
    assert "\ufeff" not in text


def test_extract_xlsx() -> None:
    from catalog.documents.extract import extract_text

    text = extract_text(str(_FIXTURES / "sample.xlsx"), "xlsx")
    assert "## Sheet: Data" in text
    assert "## Sheet: Notes" in text
    assert "| name | city | score |" in text
    assert "| Anna | Moscow | 42 |" in text
    assert "первый" in text
    assert "второй" in text


def test_extract_xlsx_escapes_pipes_and_newlines(tmp_path) -> None:
    import openpyxl
    from catalog.documents.extract import extract_text

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["a|b", "line1\nline2"])
    ws.append(["ok", "c\r\nd"])
    path = tmp_path / "special.xlsx"
    wb.save(path)
    wb.close()

    text = extract_text(str(path), "xlsx")
    assert "| a\\|b | line1 line2 |" in text
    assert "| ok | c d |" in text
    assert "| a|b |" not in text


def test_extract_xlsx_formula_without_cache(tmp_path) -> None:
    import openpyxl
    from catalog.documents.extract import extract_text

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["x", "y", "sum"])
    ws.append([2, 3, "=A2+B2"])
    path = tmp_path / "formula.xlsx"
    wb.save(path)
    wb.close()

    text = extract_text(str(path), "xlsx")
    assert "| x | y | sum |" in text
    assert "| 2 | 3 | =A2+B2 |" in text


def test_extract_xlsx_merged_header(tmp_path) -> None:
    import openpyxl
    from catalog.documents.extract import extract_text

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "Title"
    ws.merge_cells("A1:C1")
    ws["A2"] = "a"
    ws["B2"] = "b"
    ws["C2"] = "c"
    path = tmp_path / "merged.xlsx"
    wb.save(path)
    wb.close()

    text = extract_text(str(path), "xlsx")
    assert "| Title | Title | Title |" in text
    assert "| a | b | c |" in text


def test_extract_pdf() -> None:
    from catalog.documents.extract import extract_text

    text = extract_text(str(_FIXTURES / "sample-text.pdf"), "pdf")
    assert "--- page 1 ---" in text
    assert "--- page 2 ---" in text
    assert "Hello Page 1" in text
    assert "World Page 2" in text


def test_extract_pdf_scan_returns_warning() -> None:
    from catalog.documents.extract import extract_text

    text = extract_text(str(_FIXTURES / "sample-scan.pdf"), "pdf")
    assert "no extractable text" in text.lower()
    assert "scanned" in text.lower()


def test_delete_document(client, db) -> None:
    from pathlib import Path

    from catalog.storage.repo_document import get_document

    uploaded = client.post(
        "/documents", files={"file": ("note.md", b"# Title\n", "text/markdown")}
    )
    assert uploaded.status_code == 200
    doc_id = uploaded.json()["id"]
    row = get_document(db, doc_id)
    assert row is not None
    workspace = Path(client.app.state.workspace)
    assert (workspace / row.path).is_file()

    deleted = client.delete(f"/documents/{doc_id}")
    assert deleted.status_code == 204
    assert get_document(db, doc_id) is None
    assert not (workspace / row.path).exists()

    missing = client.delete(f"/documents/{doc_id}")
    assert missing.status_code == 404


def test_list_documents_does_not_auto_scan(client, db) -> None:
    from pathlib import Path

    from catalog.storage.repo_document import get_document

    kept_id = _upload(client, "keep.md", b"keep")
    orphan_id = _upload(client, "gone.md", b"gone")
    orphan = get_document(db, orphan_id)
    assert orphan is not None
    (Path(client.app.state.workspace) / orphan.path).unlink()

    listing = client.get("/documents")
    assert listing.status_code == 200
    ids = [d["id"] for d in listing.json()]
    assert kept_id in ids
    assert orphan_id in ids
    assert get_document(db, orphan_id) is not None


def test_reconcile_documents_endpoint(client, db) -> None:
    from pathlib import Path

    from catalog.storage.repo_document import get_document

    orphan_id = _upload(client, "gone.md", b"gone")
    orphan = get_document(db, orphan_id)
    assert orphan is not None
    (Path(client.app.state.workspace) / orphan.path).unlink()

    resp = client.post("/documents/reconcile")
    assert resp.status_code == 200
    assert resp.json()["removed"] == [orphan.path]
    assert get_document(db, orphan_id) is None


def test_workspaces_rescan_endpoint(client, db) -> None:
    from pathlib import Path

    from catalog.storage.repo_document import list_documents

    workspace = Path(client.app.state.workspace)
    (workspace / "nested").mkdir()
    (workspace / "nested" / "a.md").write_text("hello", encoding="utf-8")
    (workspace / "skip.exe").write_bytes(b"MZ")
    (workspace / ".hidden.md").write_text("x", encoding="utf-8")

    resp = client.post("/workspaces/rescan")
    assert resp.status_code == 200
    body = resp.json()
    assert "nested/a.md" in body["added"]
    assert any(p.endswith("skip.exe") or p == "skip.exe" for p in body["skipped"])
    indexed = {d.path for d in list_documents(db)}
    for rel_path in body["added"]:
        assert rel_path in indexed
    again = client.post("/workspaces/rescan").json()
    assert again["added"] == []
    assert again["updated"] == []
    assert again["renamed"] == []
    assert again["removed"] == []
    assert "nested/a.md" in indexed


# --------------------------------------------------------------------------- #
# Planner WebSocket
# --------------------------------------------------------------------------- #


def test_ws_session_planner(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    provider.script = [_completion("Вот план: разобрать документ по разделам.")]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text("сделай план")
        frames = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break

    types = [f["type"] for f in frames]
    assert "step" in types
    assert "token" in types
    assert frames[-1]["type"] == "finish"
    assert frames[-1]["status"] == "ok"

    token = next(f for f in frames if f["type"] == "token")
    assert "план" in token["delta"]

    # Conversation persisted: the user turn and the assistant reply.
    msgs = list_messages(db, session_id)
    roles = [m["role"] for m in msgs]
    assert roles.count("user") >= 1
    assert roles.count("assistant") >= 1


def test_workspace_idle_ws_does_not_block_switch(client, settings) -> None:
    other = Path(settings.fs_root) / "other-ws"
    other.mkdir()
    session_id = client.post("/sessions").json()["id"]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        assert ws.receive_json()["type"] == "suggestions"
        assert client.get("/workspaces/busy").json() == {
            "busy": False,
            "reason": None,
        }
        opened = client.post(
            "/workspaces/open", json={"path": str(other), "confirm": True}
        )
        assert opened.status_code == 200
        assert opened.json()["status"] == "ok"

    assert client.get("/workspaces/busy").json() == {"busy": False, "reason": None}


def test_ws_session_attach_documents_and_prompt(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    doc_a = _upload(client, "alpha.md", b"# Alpha\n")
    doc_b = _upload(client, "beta.md", b"# Beta\n")
    provider.script = [_completion("план по документам")]

    payload = {
        "type": "user",
        "content": "работай с этими",
        "doc_ids": [doc_a, doc_b],
    }
    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text(json.dumps(payload))
        frames: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break

    session_docs = next(f for f in frames if f["type"] == "session_docs")
    assert {d["id"] for d in session_docs["documents"]} == {doc_a, doc_b}
    titles = {d["title"] for d in session_docs["documents"]}
    assert "alpha" in titles
    assert "beta" in titles

    listing = client.get(f"/sessions/{session_id}/documents")
    assert listing.status_code == 200
    listed_ids = [d["id"] for d in listing.json()]
    assert set(listed_ids) == {doc_a, doc_b}

    provider.script = [_completion("ещё раз")]
    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text(json.dumps(payload))
        while True:
            frame = ws.receive_json()
            if frame.get("type") == "finish":
                break

    listing2 = client.get(f"/sessions/{session_id}/documents").json()
    assert len(listing2) == 2

    assert provider.requests
    system = provider.requests[0]["messages"][0]
    assert system.role == "system"
    assert doc_a in system.content
    assert doc_b in system.content
    assert "alpha" in system.content
    assert "beta" in system.content


def test_create_session_attaches_doc_ids_before_response(client) -> None:
    doc_id = _upload(client, "first.md", b"# First\n")
    resp = client.post("/sessions", json={"doc_ids": [doc_id, "missing-id"]})
    assert resp.status_code == 200
    body = resp.json()
    session_id = body["id"]
    assert body["skipped_doc_ids"] == ["missing-id"]
    listing = client.get(f"/sessions/{session_id}/documents")
    assert listing.status_code == 200
    assert [d["id"] for d in listing.json()] == [doc_id]


def test_get_session_documents_404(client) -> None:
    resp = client.get("/sessions/missing/documents")
    assert resp.status_code == 404


def test_detach_session_document(client, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    doc_a = _upload(client, "alpha.md", b"# Alpha\n")
    doc_b = _upload(client, "beta.md", b"# Beta\n")
    attach_documents(db, session_id, [doc_a, doc_b])

    listed = client.get(f"/sessions/{session_id}/documents").json()
    assert {d["id"] for d in listed} == {doc_a, doc_b}

    resp = client.delete(f"/sessions/{session_id}/documents/{doc_a}")
    assert resp.status_code == 204
    listed2 = client.get(f"/sessions/{session_id}/documents").json()
    assert [d["id"] for d in listed2] == [doc_b]
    assert get_document(db, doc_a) is not None

    resp2 = client.delete(f"/sessions/{session_id}/documents/{doc_a}")
    assert resp2.status_code == 404

    resp3 = client.delete(f"/sessions/missing/documents/{doc_b}")
    assert resp3.status_code == 404


def test_ws_attach_makes_document_available_to_tools(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    doc_id = _upload(client, "control.md", b"from chat control")
    other_id = _upload(client, "other.md", b"not attached")
    provider.script = [_completion("ok")]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text(
            json.dumps(
                {
                    "type": "user",
                    "content": "вот документ",
                    "doc_ids": [doc_id],
                }
            )
        )
        while True:
            frame = ws.receive_json()
            if frame.get("type") == "finish":
                break

    workspace = Path(client.app.state.workspace)
    tools = build_document_tools(db, workspace, session_id)

    async def _list():
        _, fn = tools.get("list_documents")
        return await fn()

    async def _read(doc_id_: str):
        _, fn = tools.get("read_document")
        return await fn(doc_id=doc_id_)

    listed = asyncio.run(_list())
    assert [item["id"] for item in listed] == [doc_id]
    assert asyncio.run(_read(doc_id))["text"] == "from chat control"
    assert asyncio.run(_read(other_id)) == {
        "error": "document_not_available_in_session"
    }

    global_ids = {d["id"] for d in client.get("/documents").json()}
    assert {doc_id, other_id}.issubset(global_ids)


def test_session_reopen_get_documents_restores_composition(client, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    doc_a = _upload(client, "keep-a.md", b"a")
    doc_b = _upload(client, "keep-b.md", b"b")
    attach_documents(db, session_id, [doc_a, doc_b])

    first = client.get(f"/sessions/{session_id}/documents")
    assert first.status_code == 200
    assert {d["id"] for d in first.json()} == {doc_a, doc_b}

    second = client.get(f"/sessions/{session_id}/documents")
    assert second.status_code == 200
    assert {d["id"] for d in second.json()} == {doc_a, doc_b}

    workspace = Path(client.app.state.workspace)
    tools = build_document_tools(db, workspace, session_id)

    async def _list():
        _, fn = tools.get("list_documents")
        return await fn()

    listed = asyncio.run(_list())
    assert {item["id"] for item in listed} == {doc_a, doc_b}


def test_parse_suggestions_extracts_and_strips() -> None:
    """parse_suggestions pulls items out and removes the block (CATALOG-13)."""
    from catalog.api.sessions import parse_suggestions

    text = "Вот план.\n\n<suggestions>Шаг 1 | Шаг 2 | Шаг 3</suggestions>"
    clean, items = parse_suggestions(text)
    assert clean == "Вот план."
    assert items == ["Шаг 1", "Шаг 2", "Шаг 3"]


def test_parse_suggestions_no_block_unchanged() -> None:
    """Without a block the text is returned as-is with an empty list."""
    from catalog.api.sessions import parse_suggestions

    text = "Просто ответ без подсказок."
    clean, items = parse_suggestions(text)
    assert clean == text
    assert items == []


def test_parse_suggestions_empty_items() -> None:
    """A block with only separators yields no items and is still stripped."""
    from catalog.api.sessions import parse_suggestions

    clean, items = parse_suggestions("Ответ.\n<suggestions>  |  |  </suggestions>")
    assert items == []
    assert "<suggestions>" not in clean
    assert clean == "Ответ."


def test_ws_session_starter_suggestions(client, provider, db) -> None:
    """An empty session receives a starter suggestions frame right after accept (CATALOG-13)."""
    session_id = client.post("/sessions").json()["id"]
    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        frame = ws.receive_json()
    assert frame["type"] == "suggestions"
    assert isinstance(frame["items"], list)
    assert len(frame["items"]) >= 1


def test_ws_session_emits_suggestions_frame(client, provider, db) -> None:
    """A model reply with a <suggestions> block yields a suggestions frame and is stripped (CATALOG-13)."""
    session_id = client.post("/sessions").json()["id"]
    provider.script = [
        _completion("Вот план.\n<suggestions>Изучи документы | Опиши задачу</suggestions>")
    ]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text("сделай план")
        frames: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break

    sug_frames = [f for f in frames if f["type"] == "suggestions"]
    # The starter frame (empty session) + the model frame.
    assert len(sug_frames) >= 1
    model_sug = sug_frames[-1]
    assert model_sug["items"] == ["Изучи документы", "Опиши задачу"]

    token = next(f for f in frames if f["type"] == "token")
    assert "<suggestions>" not in token["delta"]
    assert "Вот план" in token["delta"]

    # Persisted assistant text is the cleaned one.
    msgs = list_messages(db, session_id)
    assistant = [m for m in msgs if m["role"] == "assistant"][-1]
    assert "<suggestions>" not in assistant["content"]


def test_planner_uses_active_model(client, provider, db) -> None:
    """Changing the model via POST /settings drives the planner LLM call (CATALOG-14)."""
    session_id = client.post("/sessions").json()["id"]
    client.post("/settings", json={"model": "glm-4.6"})
    provider.script = [_completion("план")]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text("сделай план")
        while True:
            frame = ws.receive_json()
            if frame.get("type") == "finish":
                break

    assert provider.requests, "planner did not call the provider"
    assert provider.requests[0]["model"] == "glm-4.6"


def test_ws_session_idle_keepalive(client, monkeypatch) -> None:
    """Idle planner WS emits ping frames before the typical ~5 min proxy timeout (CATALOG-23)."""
    import catalog.api.sessions as sessions_mod

    monkeypatch.setattr(sessions_mod, "WS_KEEPALIVE_INTERVAL_S", 0.05)
    session_id = client.post("/sessions").json()["id"]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        assert ws.receive_json()["type"] == "suggestions"
        ping = ws.receive_json()
        assert ping == {"type": "ping"}
        ws.send_text('{"type":"pong"}')
        ping2 = ws.receive_json()
        assert ping2 == {"type": "ping"}


def test_ws_session_cancel(client, provider, db) -> None:
    """Cancelling a running planner turn sends finish{cancelled} and keeps the session alive (CATALOG-11)."""
    from tests.conftest import FakeProvider as _ConfFakeProvider

    class _BlockingProvider(_ConfFakeProvider):
        """Provider whose complete() blocks until cancelled."""

        def __init__(self) -> None:
            super().__init__(script=[])
            self.was_cancelled = False

        async def complete(self, *args, **kwargs):  # type: ignore[override]
            import asyncio

            try:
                await asyncio.Event().wait()
            except asyncio.CancelledError:
                self.was_cancelled = True
                raise

    blocking = _BlockingProvider()
    client.app.state.provider = blocking

    session_id = client.post("/sessions").json()["id"]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        # Send a message, then immediately a cancel frame. The cancel is read
        # by the concurrent listener while the agent is blocked in complete().
        ws.send_text("долгий вопрос")
        ws.send_text('{"type":"cancel"}')

        frames: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break
        assert client.get("/workspaces/busy").json() == {"busy": False, "reason": None}

    finish = frames[-1]
    assert finish["type"] == "finish"
    assert finish["status"] == "cancelled"
    # The provider observed the cancellation at its await point.
    assert blocking.was_cancelled is True
    assert client.app.state.active_planner_turns == 0

    # The session is still alive: a follow-up message completes normally.
    provider.script = [_completion("готово")]
    client.app.state.provider = provider
    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        ws.send_text("ещё вопрос")
        frames2: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames2.append(frame)
            if frame.get("type") == "finish":
                break
    assert frames2[-1]["status"] == "ok"


def test_wait_work_or_ws_prefers_done_work_over_cancel() -> None:
    from catalog.api.sessions import _wait_work_or_ws

    class _DummyWS:
        async def send_json(self, data):
            raise AssertionError("should not ping")

    async def _run() -> None:
        work = asyncio.create_task(asyncio.sleep(0))
        receive = asyncio.create_task(asyncio.sleep(0, result='{"type":"cancel"}'))
        await work
        await receive
        kind, payload = await _wait_work_or_ws(_DummyWS(), work, receive)
        assert kind == "work"
        assert payload is None

    asyncio.run(_run())


def test_wait_work_or_ws_cancel_while_work_running() -> None:
    from catalog.api.sessions import _wait_work_or_ws

    class _DummyWS:
        async def send_json(self, data):
            raise AssertionError("should not ping")

    async def _run() -> None:
        blocker = asyncio.Event()
        work = asyncio.create_task(blocker.wait())
        receive = asyncio.create_task(asyncio.sleep(0, result='{"type":"cancel"}'))
        kind, payload = await _wait_work_or_ws(_DummyWS(), work, receive)
        assert kind == "cancel"
        assert payload == '{"type":"cancel"}'
        work.cancel()
        try:
            await work
        except asyncio.CancelledError:
            pass

    asyncio.run(_run())


def test_ws_session_idle_cancel_sends_noop(client) -> None:
    session_id = client.post("/sessions").json()["id"]
    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        assert ws.receive_json()["type"] == "suggestions"
        ws.send_text('{"type":"cancel"}')
        frame = ws.receive_json()
        assert frame == {"type": "finish", "status": "noop"}


def test_ws_session_planner_keepalive_during_turn(client, db, monkeypatch) -> None:
    import catalog.api.sessions as sessions_mod
    from tests.conftest import HoldCompleteProvider

    monkeypatch.setattr(sessions_mod, "WS_KEEPALIVE_INTERVAL_S", 0.05)
    hold = HoldCompleteProvider(_completion("готово"))
    client.app.state.provider = hold
    session_id = client.post("/sessions").json()["id"]

    with client.websocket_connect(f"/sessions/{session_id}") as ws:
        assert ws.receive_json()["type"] == "suggestions"
        ws.send_text("вопрос")
        frames: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "ping":
                break
        hold.release.set()
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break

    assert any(f.get("type") == "ping" for f in frames)
    assert frames[-1]["status"] == "ok"


# --------------------------------------------------------------------------- #
# Skill build / commit / list
# --------------------------------------------------------------------------- #


def test_build_skill_from_session(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="Хочу скилл-саммаризатор.")

    provider.script = [_completion(tool_calls=[_build_skill_call(name="Summarizer")])]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    skill_id = body["skill_id"]
    # CATALOG-6: build returns a preview config for the settings modal.
    assert body["config"]["name"] == "Summarizer"
    assert body["config"]["model"] == "test/model"

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.status == "draft"
    assert skill.config.name == "Summarizer"
    assert skill.config.allowed_tools == ["read_document"]
    assert [vc.check for vc in skill.config.verify_checks] == ["non_empty"]


def test_edit_skill_starts_session_with_skill_id(client, provider, db) -> None:
    """POST /skills/{id}/edit creates a session linked to the skill (CATALOG-17)."""
    skill_id = _seed_committed_skill(db, name="Original")

    resp = client.post(f"/skills/{skill_id}/edit")
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["skill_id"] == skill_id
    session_id = body["session_id"]

    session = get_session(db, session_id)
    assert session is not None
    assert session.skill_id == skill_id

    msgs = list_messages(db, session_id)
    assert len(msgs) == 1
    seed = msgs[0]["content"]
    assert "Original" in seed
    assert skill_id in seed
    assert "set_skill_meta" in seed
    assert "save_skill_prompt" in seed
    assert "build_skill" not in seed
    assert "You summarize the document." not in seed

    arts = client.get(f"/sessions/{session_id}/artifacts").json()
    types = {a["type"] for a in arts}
    assert "meta" in types
    assert "prompt" in types


def test_edit_skill_missing_returns_404(client, db) -> None:
    resp = client.post("/skills/does-not-exist/edit")
    assert resp.status_code == 404


def test_build_from_edit_session_updates_same_skill_and_drops_to_draft(
    client, provider, db
) -> None:
    """Building from an edit session updates the same skill (CATALOG-17).

    A committed skill drops back to draft after the edit is saved.
    Edit sessions seed artifacts; build packs them without LLM (CATALOG-53).
    """
    skill_id = _seed_committed_skill(db, name="Original")

    session_id = client.post(f"/skills/{skill_id}/edit").json()["session_id"]
    client.patch(
        f"/sessions/{session_id}/skill-meta",
        json={
            "name": "Renamed",
            "description": "updated",
            "kind": "agent",
            "allowed_tools": ["read_document"],
            "verify_checks": [{"check": "non_empty"}],
        },
    )
    client.patch(
        f"/sessions/{session_id}/artifacts/prompt",
        json={"content": "You process the document as instructed."},
    )
    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text
    assert resp.json()["skill_id"] == skill_id  # same id, not a new skill

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.id == skill_id
    assert skill.name == "Renamed"
    assert skill.config.name == "Renamed"
    assert skill.status == "draft"  # committed -> draft after save


def test_build_from_edit_session_draft_stays_draft(client, provider, db) -> None:
    """Editing a draft skill keeps it a draft (no forced status change)."""
    session0 = client.post("/sessions").json()["id"]
    add_message(db, session_id=session0, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="S")])]
    skill_id = client.post(f"/sessions/{session0}/skills").json()["skill_id"]
    assert get_skill(db, skill_id).status == "draft"

    session_id = client.post(f"/skills/{skill_id}/edit").json()["session_id"]
    client.patch(
        f"/sessions/{session_id}/skill-meta",
        json={
            "name": "S2",
            "description": "Skill built from a planning session.",
            "kind": "agent",
            "allowed_tools": ["read_document"],
            "verify_checks": [{"check": "non_empty"}],
        },
    )
    client.patch(
        f"/sessions/{session_id}/artifacts/prompt",
        json={"content": "You process the document as instructed."},
    )
    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.status == "draft"
    assert skill.name == "S2"


def test_build_without_skill_id_still_creates_new_skill(client, provider, db) -> None:
    """Regression: a regular (non-edit) session still creates a new skill."""
    existing_id = _seed_committed_skill(db, name="Existing")

    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="Brand new")])]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text
    new_skill_id = resp.json()["skill_id"]

    assert new_skill_id != existing_id
    new_skill = get_skill(db, new_skill_id)
    assert new_skill is not None
    assert new_skill.name == "Brand new"
    # The pre-existing skill is untouched.
    existing = get_skill(db, existing_id)
    assert existing is not None
    assert existing.name == "Existing"


def test_configure_skill_updates_model_provider_reasoning(client, provider, db) -> None:
    """PATCH /skills/{id}/configure overrides model/provider/reasoning (CATALOG-6)."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="S")])]

    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]

    resp = client.patch(
        f"/skills/{skill_id}/configure",
        json={"model": "glm-4.6", "provider": "zai", "reasoning": "high"},
    )
    assert resp.status_code == 200, resp.text
    cfg = resp.json()["config"]
    assert cfg["model"] == "glm-4.6"
    assert cfg["provider"] == "zai"
    assert cfg["reasoning"] == "high"

    # Persisted in config_json.
    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.model == "glm-4.6"
    assert skill.config.provider == "zai"
    assert skill.config.reasoning == "high"


def test_configure_skill_saves_input_arity(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="S")])]
    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]

    resp = client.patch(f"/skills/{skill_id}/configure", json={"input_arity": 2})
    assert resp.status_code == 200, resp.text
    assert resp.json()["config"]["input_arity"] == 2
    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.input_arity == 2

    resp = client.patch(f"/skills/{skill_id}/configure", json={"input_arity": 1})
    assert resp.status_code == 200, resp.text
    assert resp.json()["config"]["input_arity"] == 1
    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.input_arity == 1

    resp = client.patch(f"/skills/{skill_id}/configure", json={"input_arity": None})
    assert resp.status_code == 200, resp.text
    assert resp.json()["config"]["input_arity"] is None
    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.input_arity is None

    resp = client.patch(f"/skills/{skill_id}/configure", json={"model": "other"})
    assert resp.status_code == 200, resp.text
    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.input_arity is None
    assert skill.config.model == "other"

    resp = client.patch(f"/skills/{skill_id}/configure", json={"input_arity": 3})
    assert resp.status_code == 422


def test_configure_skill_requires_draft(client, provider, db) -> None:
    """Configure is rejected (409) once the skill is committed."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="S")])]
    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]
    update_status(db, skill_id, "committed")

    resp = client.patch(f"/skills/{skill_id}/configure", json={"model": "other"})
    assert resp.status_code == 409


def test_configure_skill_renames_draft(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="Old")])]
    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]

    resp = client.patch(
        f"/skills/{skill_id}/configure",
        json={"name": "New Name", "model": "glm-4.6"},
    )
    assert resp.status_code == 200, resp.text
    assert resp.json()["config"]["name"] == "New Name"
    assert resp.json()["config"]["model"] == "glm-4.6"

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.name == "New Name"
    assert skill.config.name == "New Name"
    assert skill.config.model == "glm-4.6"

    listed = client.get("/skills").json()
    match = next(s for s in listed if s["id"] == skill_id)
    assert match["name"] == "New Name"


def test_rename_committed_skill(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="Old")])]
    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]
    update_status(db, skill_id, "committed")

    resp = client.patch(f"/skills/{skill_id}", json={"name": "Renamed"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["id"] == skill_id
    assert body["name"] == "Renamed"
    assert body["status"] == "committed"

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.name == "Renamed"
    assert skill.config.name == "Renamed"

    listed = client.get("/skills").json()
    match = next(s for s in listed if s["id"] == skill_id)
    assert match["name"] == "Renamed"


def test_rename_skill_rejects_empty_name(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    provider.script = [_completion(tool_calls=[_build_skill_call(name="S")])]
    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]

    resp = client.patch(f"/skills/{skill_id}", json={"name": "   "})
    assert resp.status_code == 422

    resp = client.patch(f"/skills/{skill_id}/configure", json={"name": ""})
    assert resp.status_code == 422


def test_list_models_endpoint(client, provider, monkeypatch) -> None:
    """GET /models returns the active provider catalog with reasoning info."""
    from catalog.llm.base import ModelInfo

    async def _models() -> list[ModelInfo]:
        return [
            ModelInfo(
                id="glm-4.6",
                name="GLM-4.6",
                context_length=131072,
                supports_reasoning=True,
                reasoning_variants=["low", "medium", "high"],
            )
        ]

    monkeypatch.setattr(provider, "list_models", _models)
    resp = client.get("/models")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert len(data) == 1
    assert data[0]["id"] == "glm-4.6"
    assert data[0]["supports_reasoning"] is True
    assert data[0]["reasoning_variants"] == ["low", "medium", "high"]


def test_list_providers_endpoint(client) -> None:
    """GET /providers returns at least one provider, the active one flagged."""
    resp = client.get("/providers")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert isinstance(data, list)
    assert len(data) >= 1
    assert any(p["active"] for p in data)


def test_provider_for_skill_resolves_pinned_provider() -> None:
    """A skill's pinned provider is used at apply; unknown/empty falls back (CATALOG-6)."""
    from catalog.llm.factory import provider_for_skill

    active = object()
    zai = object()
    providers = {"openrouter": object(), "zai": zai}

    assert provider_for_skill(providers, active, "zai") is zai
    # Empty provider name -> active provider.
    assert provider_for_skill(providers, active, "") is active
    # Unknown provider name -> active provider (graceful fallback).
    assert provider_for_skill(providers, active, "nope") is active
    # No providers dict -> active provider.
    assert provider_for_skill(None, active, "zai") is active


def test_build_persists_provider_and_reasoning(client, provider, db) -> None:
    """Build carries provider/reasoning from the model's tool args (CATALOG-6)."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")
    call = _build_skill_call(name="S")
    call.arguments["provider"] = "zai"
    call.arguments["reasoning"] = "high"
    provider.script = [_completion(tool_calls=[call])]

    skill_id = client.post(f"/sessions/{session_id}/skills").json()["skill_id"]
    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.provider == "zai"
    assert skill.config.reasoning == "high"


def test_get_and_update_settings(client) -> None:
    """GET/POST /settings read and switch the runtime model (CATALOG-14)."""
    before = client.get("/settings").json()
    assert before["model"] == "test/model"

    resp = client.post("/settings", json={"model": "glm-4.6"})
    assert resp.status_code == 200, resp.text
    assert resp.json()["model"] == "glm-4.6"
    # Persisted in app state.
    assert client.app.state.active_model == "glm-4.6"
    assert client.get("/settings").json()["model"] == "glm-4.6"


def test_update_settings_unknown_provider_404(client) -> None:
    """Switching to an unconfigured provider is rejected (CATALOG-14)."""
    resp = client.post("/settings", json={"provider": "does-not-exist"})
    assert resp.status_code == 404


def test_update_settings_switches_active_provider(client) -> None:
    """POST /settings provider=openrouter resolves the active instance (CATALOG-14)."""
    providers = client.app.state.providers
    assert "openrouter" in providers
    client.app.state.active_model = "glm-4.6"
    resp = client.post("/settings", json={"provider": "openrouter"})
    assert resp.status_code == 200, resp.text
    assert client.app.state.provider is providers["openrouter"]
    assert client.app.state.active_model == "test/model"
    assert resp.json()["model"] == "test/model"


def test_provider_models_endpoint(client, monkeypatch) -> None:
    """GET /providers/{id}/models lists a specific provider's catalog (CATALOG-14)."""
    from catalog.llm.base import ModelInfo

    class _FakeProv:
        async def list_models(self) -> list[ModelInfo]:
            return [ModelInfo(id="z-glm", name="Z-GLM", context_length=8192)]

    client.app.state.providers = {**client.app.state.providers, "zfake": _FakeProv()}
    resp = client.get("/providers/zfake/models")
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert len(data) == 1
    assert data[0]["id"] == "z-glm"

    # Unknown provider -> 404.
    assert client.get("/providers/nope/models").status_code == 404


def test_build_skill_invalid_allowed_tools_returns_422(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")

    bad = _completion(
        tool_calls=[_build_skill_call(allowed_tools=["nonexistent_tool"], verify_checks=[])]
    )
    provider.script = [bad, bad, bad]  # 1 initial + 2 retries, all invalid

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 422
    detail = resp.json()["detail"]
    assert "failed to build a valid skill after retries" in detail
    assert "session LLM timeout" in detail


def test_build_agent_skill_empty_allowed_tools_adds_read_document(
    client, provider, db
) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")

    provider.script = [
        _completion(tool_calls=[_build_skill_call(name="NoTools", allowed_tools=[])])
    ]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text
    skill = get_skill(db, resp.json()["skill_id"])
    assert skill is not None
    assert skill.config.kind == "agent"
    assert "read_document" in skill.config.allowed_tools


def test_build_skill_timeout_returns_504(client, provider, db) -> None:
    from catalog.llm.timeout import LLMTimeoutError

    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")

    async def _timeout(*_args, **_kwargs):
        raise LLMTimeoutError(
            "openrouter request timed out after 60s (3 retries exhausted)",
            timeout_seconds=60,
        )

    provider.complete = _timeout  # type: ignore[method-assign]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 504
    detail = resp.json()["detail"]
    assert "timed out after 60s" in detail
    assert "Increase the session LLM timeout" in detail


def test_build_skill_provider_runtime_error_not_timeout_advice(
    client, provider, db
) -> None:
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="make a skill")

    async def _rate_limited(*_args, **_kwargs):
        raise RuntimeError("OpenRouter returned HTTP 429 after 3 retries")

    provider.complete = _rate_limited  # type: ignore[method-assign]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 502
    detail = resp.json()["detail"]
    assert "429" in detail
    assert "session LLM timeout" not in detail.lower()


def test_session_timeout_default_and_patch(client) -> None:
    session_id = client.post("/sessions").json()["id"]

    got = client.get(f"/sessions/{session_id}")
    assert got.status_code == 200
    assert got.json()["llm_timeout_seconds"] == 60

    listed = client.get("/sessions").json()
    by_id = {s["id"]: s for s in listed}
    assert by_id[session_id]["llm_timeout_seconds"] == 60

    patched = client.patch(
        f"/sessions/{session_id}",
        json={"llm_timeout_seconds": 120},
    )
    assert patched.status_code == 200, patched.text
    assert patched.json()["llm_timeout_seconds"] == 120
    assert client.get(f"/sessions/{session_id}").json()["llm_timeout_seconds"] == 120

    too_low = client.patch(
        f"/sessions/{session_id}",
        json={"llm_timeout_seconds": 10},
    )
    assert too_low.status_code == 422

    too_high = client.patch(
        f"/sessions/{session_id}",
        json={"llm_timeout_seconds": 999},
    )
    assert too_high.status_code == 422

    missing = client.patch(
        "/sessions/does-not-exist",
        json={"llm_timeout_seconds": 90},
    )
    assert missing.status_code == 404


def test_build_script_skill_valid_code(client, provider, db) -> None:
    """A kind=script skill with valid Python code is built as draft."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="uppercase the doc")
    code = "result = document.upper()\n"
    upsert_script_dry_run(
        db,
        session_id=session_id,
        slot="script",
        sha256=code_sha256(code),
        ok=True,
        stage="run",
    )

    provider.script = [
        _completion(
            tool_calls=[_build_script_skill_call(code=code)]
        )
    ]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text
    skill_id = resp.json()["skill_id"]

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.status == "draft"
    assert skill.config.kind == "script"
    assert skill.config.code == "result = document.upper()\n"
    # Scripts have no tools.
    assert skill.config.allowed_tools == []


def test_build_script_skill_forbidden_import_returns_422(client, provider, db) -> None:
    """A kind=script skill with a forbidden import is rejected (422)."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="read the filesystem")

    bad = _completion(
        tool_calls=[_build_script_skill_call(code="import os\nresult = 'x'\n")]
    )
    provider.script = [bad, bad, bad]  # all attempts invalid

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 422


def test_build_script_skill_dangerous_call_returns_422(client, provider, db) -> None:
    """A kind=script skill with eval() is rejected (422)."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="eval stuff")

    bad = _completion(
        tool_calls=[_build_script_skill_call(code="result = eval('1')\n")]
    )
    provider.script = [bad, bad, bad]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 422


def test_build_agent_skill_with_non_determinism_reason(client, provider, db) -> None:
    """When the task is not deterministic the model chooses kind=agent with a reason."""
    session_id = client.post("/sessions").json()["id"]
    add_message(db, session_id=session_id, role="user", content="summarize creatively")

    provider.script = [
        _completion(
            tool_calls=[
                ToolCall(
                    id="build-1",
                    name="build_skill",
                    arguments={
                        "name": "CreativeSummarizer",
                        "description": "Creative summary.",
                        "kind": "agent",
                        "non_determinism_reason": "Needs subjective judgment on tone.",
                        "system_prompt": "You summarize creatively.",
                        "allowed_tools": ["read_document"],
                        "model": "test/model",
                    },
                )
            ]
        )
    ]

    resp = client.post(f"/sessions/{session_id}/skills")
    assert resp.status_code == 200, resp.text
    skill_id = resp.json()["skill_id"]

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.config.kind == "agent"


def test_commit_skill(client, db) -> None:
    skill_id = _seed_committed_skill(db, name="Committed")
    # Seed as draft to exercise the draft -> committed transition.
    update_status(db, skill_id, "draft")

    resp = client.post(f"/skills/{skill_id}/commit")
    assert resp.status_code == 200
    assert resp.json() == {"id": skill_id, "status": "committed"}

    skill = get_skill(db, skill_id)
    assert skill is not None
    assert skill.status == "committed"


def test_delete_draft_skill(client, db) -> None:
    skill_id = create_skill(
        db,
        name="DraftToDelete",
        description="draft",
        config=SkillConfig(
            name="DraftToDelete",
            description="draft",
            system_prompt="x",
            allowed_tools=["read_document"],
            model="test/model",
        ),
        status="draft",
    )

    resp = client.delete(f"/skills/{skill_id}")
    assert resp.status_code == 204
    assert get_skill(db, skill_id) is None
    ids = [r["id"] for r in client.get("/skills").json()]
    assert skill_id not in ids


def test_delete_committed_skill_cascades_runs(client, db) -> None:
    from catalog.skills.repo_run import create_run, get_run

    skill_id = _seed_committed_skill(db, name="CommittedToDelete")
    doc_id = _upload(client, "in.md", b"body")
    run_id = create_run(
        db, skill_id=skill_id, session_id=None, input_doc_ids=[doc_id]
    )

    resp = client.delete(f"/skills/{skill_id}")
    assert resp.status_code == 204
    assert get_skill(db, skill_id) is None
    assert get_run(db, run_id) is None
    ids = [r["id"] for r in client.get("/skills").json()]
    assert skill_id not in ids


def test_delete_skill_missing_returns_404(client) -> None:
    resp = client.delete("/skills/does-not-exist")
    assert resp.status_code == 404


def test_list_skills(client, db) -> None:
    first = _seed_committed_skill(db, name="First")
    _seed_committed_skill(db, name="Second")

    resp = client.get("/skills")
    assert resp.status_code == 200
    rows = resp.json()
    ids = [r["id"] for r in rows]
    assert first in ids
    assert len(ids) >= 2

    # Optional status filter.
    draft_resp = client.get("/skills?status=draft")
    assert draft_resp.status_code == 200
    assert all(r["status"] == "draft" for r in draft_resp.json())


# --------------------------------------------------------------------------- #
# Capability tags (CATALOG-8)
# --------------------------------------------------------------------------- #


def test_compute_tags_agent_script_mixed_legacy() -> None:
    """compute_tags derives python/ai from the config (CATALOG-8 rules)."""
    # Pure agent: prompt, no code -> ["ai"].
    agent = SkillConfig(
        name="a",
        description="d",
        system_prompt="You do the task.",
        allowed_tools=[],
        model="test/model",
        kind="agent",
    )
    assert compute_tags(agent) == ["ai"]

    # Pure script: code, no prompt -> ["python"].
    script = SkillConfig(
        name="s",
        description="d",
        system_prompt="",
        allowed_tools=[],
        model="test/model",
        kind="script",
        code="result = document.upper()\n",
    )
    assert compute_tags(script) == ["python"]

    # Mixed: agent kind that also carries code -> both tags.
    mixed = SkillConfig(
        name="m",
        description="d",
        system_prompt="You do the task.",
        allowed_tools=[],
        model="test/model",
        kind="agent",
        code="result = 1\n",
    )
    assert compute_tags(mixed) == ["python", "ai"]

    # Legacy: no kind -> defaults to "agent" -> ["ai"].
    legacy = SkillConfig(
        name="l",
        description="d",
        system_prompt="You do the task.",
        allowed_tools=[],
        model="test/model",
    )
    assert legacy.kind == "agent"
    assert compute_tags(legacy) == ["ai"]


def test_compute_tags_pipeline() -> None:
    from catalog.skills.config import PipelineStep

    pipe = SkillConfig(
        name="p",
        description="d",
        system_prompt="",
        allowed_tools=[],
        model="test/model",
        kind="pipeline",
        steps=[
            PipelineStep(id="a", type="script", input="documents", code="result = document\n"),
            PipelineStep(
                id="b",
                type="llm",
                input="previous",
                system_prompt="rewrite",
            ),
        ],
    )
    assert compute_tags(pipe) == ["python", "ai"]


def test_list_skills_endpoint_returns_tags(client, db) -> None:
    """GET /skills surfaces computed tags per skill (CATALOG-8)."""
    # An agent skill (the default helper) and a script skill.
    _seed_committed_skill(db, name="AgentSkill")
    script_config = SkillConfig(
        name="ScriptSkill",
        description="deterministic script",
        system_prompt="",
        allowed_tools=[],
        model="test/model",
        kind="script",
        code="result = document.upper()\n",
    )
    create_skill(
        db,
        name=script_config.name,
        description=script_config.description,
        config=script_config,
        status="committed",
    )

    rows = client.get("/skills").json()
    by_name = {r["name"]: r for r in rows}
    assert by_name["AgentSkill"]["tags"] == ["ai"]
    assert by_name["ScriptSkill"]["tags"] == ["python"]


def test_list_skills_returns_input_arity(client, db) -> None:
    config = SkillConfig(
        name="PairMerger",
        description="needs two docs",
        system_prompt="Merge.",
        allowed_tools=["read_document"],
        model="test/model",
        input_arity=2,
    )
    create_skill(
        db,
        name=config.name,
        description=config.description,
        config=config,
        status="committed",
    )
    _seed_committed_skill(db, name="AnyList")

    rows = client.get("/skills").json()
    by_name = {r["name"]: r for r in rows}
    assert by_name["PairMerger"]["input_arity"] == 2
    assert by_name["AnyList"]["input_arity"] is None


def test_list_skills_returns_model_params(client, db) -> None:
    configured = SkillConfig(
        name="Configured",
        description="has model params",
        system_prompt="Do it.",
        allowed_tools=["read_document"],
        model="glm-4.6",
        provider="zai",
        reasoning="high",
    )
    create_skill(
        db,
        name=configured.name,
        description=configured.description,
        config=configured,
        status="committed",
    )
    _seed_committed_skill(db, name="Defaults")

    rows = client.get("/skills").json()
    by_name = {r["name"]: r for r in rows}
    assert by_name["Configured"]["provider"] == "zai"
    assert by_name["Configured"]["model"] == "glm-4.6"
    assert by_name["Configured"]["reasoning"] == "high"
    assert by_name["Defaults"]["model"] == "test/model"
    assert by_name["Defaults"]["provider"] is None
    assert by_name["Defaults"]["reasoning"] is None


# --------------------------------------------------------------------------- #
# Apply (run create / stream / get)
# --------------------------------------------------------------------------- #


def _drain_run_ws(client, run_id: str) -> list[dict]:
    with client.websocket_connect(f"/runs/{run_id}/stream") as ws:
        frames: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break
    return frames


def test_apply_skill_run(client, provider, db) -> None:
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nGreat document.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]

    frames = _drain_run_ws(client, run_id)
    finish = frames[-1]
    assert finish["type"] == "finish"
    assert finish["status"] == "ok"
    assert finish["output_doc_id"] is not None

    # A verify frame was emitted and passed.
    verifies = [f for f in frames if f["type"] == "verify"]
    assert verifies and verifies[-1]["passed"] is True
    assert verifies[-1]["checks"][0]["check"] == "non_empty"
    assert verifies[-1]["checks"][0]["passed"] is True

    # GET /runs/{id} returns the persisted trace.
    run = client.get(f"/runs/{run_id}").json()
    assert run["status"] == "ok"
    assert run["output_doc_id"] == finish["output_doc_id"]
    assert run["parent_run_id"] is None
    assert run["trace"] is not None
    saved_verify = [e for e in run["trace"] if e.get("kind") == "verify"]
    assert saved_verify
    assert saved_verify[-1]["data"]["passed"] is True
    assert saved_verify[-1]["data"]["failures"] == []
    assert saved_verify[-1]["data"]["checks"][0]["check"] == "non_empty"


def test_apply_stream_preview_token_before_verify(client, provider, db) -> None:
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nGreat document.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]

    frames = _drain_run_ws(client, run_id)
    types = [f["type"] for f in frames]
    assert "token" in types
    assert "verify" in types
    assert types.index("token") < types.index("verify")
    token = next(f for f in frames if f["type"] == "token")
    assert token["delta"] == "# Result\n\nGreat document."
    assert frames[-1]["status"] == "ok"


def test_apply_stream_keepalive_during_run(client, db, monkeypatch) -> None:
    import catalog.api.sessions as sessions_mod
    from tests.conftest import HoldCompleteProvider

    monkeypatch.setattr(sessions_mod, "WS_KEEPALIVE_INTERVAL_S", 0.05)
    hold = HoldCompleteProvider(_completion("# Result\n\nGreat document."))
    client.app.state.provider = hold

    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]

    with client.websocket_connect(f"/runs/{run_id}/stream") as ws:
        frames: list[dict] = []
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "ping":
                break
        hold.release.set()
        while True:
            frame = ws.receive_json()
            frames.append(frame)
            if frame.get("type") == "finish":
                break

    assert any(f.get("type") == "ping" for f in frames)
    assert frames[-1]["type"] == "finish"
    assert frames[-1]["status"] == "ok"


def test_apply_creates_pending_run_before_stream(client, db) -> None:
    from catalog.skills.repo_run import get_run, has_running_runs

    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(db)
    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]
    row = get_run(db, run_id)
    assert row is not None
    assert row["status"] == "pending"
    assert has_running_runs(db) is True


def test_apply_stream_rejects_second_claim(client, provider, db) -> None:
    from catalog.skills.repo_run import claim_run, get_run

    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nGreat document.")]
    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]
    assert claim_run(db, run_id) is True
    assert get_run(db, run_id)["status"] == "running"

    with client.websocket_connect(f"/runs/{run_id}/stream") as ws:
        frame = ws.receive_json()
        assert frame["type"] == "error"
        assert "already in progress" in frame["message"]


def test_apply_stream_validation_failure_cancels_pending(client, db) -> None:
    from catalog.skills.repo_run import get_run, has_running_runs

    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(db)
    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]
    with db.connect() as conn:
        conn.execute("DELETE FROM skill WHERE id = ?", (skill_id,))

    with client.websocket_connect(f"/runs/{run_id}/stream") as ws:
        frame = ws.receive_json()
        assert frame["type"] == "error"
        assert frame["message"] == "skill not found"

    row = get_run(db, run_id)
    assert row is not None
    assert row["status"] == "cancelled"
    assert has_running_runs(db) is False


def test_apply_agent_prompt_reaches_llm(client, provider, db) -> None:
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nClarified.")]
    clarification = "Сфокусируйся на рисках."

    run_id = client.post(
        f"/skills/{skill_id}/apply",
        json={"doc_id": doc_id, "prompt": clarification},
    ).json()["run_id"]

    frames = _drain_run_ws(client, run_id)
    assert frames[-1]["status"] == "ok"
    assert provider.requests
    user_contents = [
        m.content
        for m in provider.requests[0]["messages"]
        if m.role == "user"
    ]
    assert any(clarification in (c or "") for c in user_contents)
    system_contents = [
        m.content
        for m in provider.requests[0]["messages"]
        if m.role == "system"
    ]
    assert system_contents == ["You summarize the document."]


def test_apply_skill_failed(client, provider, db) -> None:
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db,
        verify_checks=[VerifyCheck("has_section", params={"heading": "Missing"})],
        max_retries=2,
    )
    # 1 initial + 2 retries, none satisfy the verify check.
    provider.script = [_completion("plain text without the heading")] * 3

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]

    frames = _drain_run_ws(client, run_id)
    finish = frames[-1]
    assert finish["type"] == "finish"
    assert finish["status"] == "failed"
    assert finish["output_doc_id"] is None

    # The run row records the failure with a trace.
    run = client.get(f"/runs/{run_id}").json()
    assert run["status"] == "failed"
    assert run["trace"] is not None


def test_apply_requires_committed_skill(client, db) -> None:
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(db, name="Draft")
    update_status(db, skill_id, "draft")

    resp = client.post(f"/skills/{skill_id}/apply", json={"doc_id": doc_id})
    assert resp.status_code == 409


def test_apply_multi_doc_via_api(client, provider, db) -> None:
    """POST /skills/{id}/apply accepts doc_ids (list); GET /runs returns the list."""
    doc_a = _upload(client, "a.md", b"first source")
    doc_b = _upload(client, "b.md", b"second source")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nMulti-doc output.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_ids": [doc_a, doc_b]}
    ).json()["run_id"]

    frames = _drain_run_ws(client, run_id)
    assert frames[-1]["type"] == "finish"
    assert frames[-1]["status"] == "ok"

    run = client.get(f"/runs/{run_id}").json()
    assert run["status"] == "ok"
    assert run["input_doc_ids"] == [doc_a, doc_b]


def test_apply_preview_mode_skips_document_creation(client, provider, db) -> None:
    """persist=False ("на экран", CATALOG-18) does not create a result_md doc.

    The finish frame and GET /runs/{id} still carry the full result_text so
    the UI can render it without a document.
    """
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nOn-screen only.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id, "persist": False}
    ).json()["run_id"]

    frames = _drain_run_ws(client, run_id)
    finish = frames[-1]
    assert finish["type"] == "finish"
    assert finish["status"] == "ok"
    assert finish["output_doc_id"] is None
    assert finish["output_doc_ids"] == []
    assert finish["result_text"] == "# Result\n\nOn-screen only."
    assert finish["result_artifacts"] == {}

    run = client.get(f"/runs/{run_id}").json()
    assert run["status"] == "ok"
    assert run["output_doc_id"] is None
    assert run["result_text"] == "# Result\n\nOn-screen only."

    # No result_md document appeared in the list.
    docs_before = client.get("/documents").json()
    assert all(d["kind"] != "result_md" for d in docs_before)


def test_save_run_result_materializes_preview_into_document(
    client, provider, db
) -> None:
    """POST /runs/{id}/save creates a document from a preview run's text."""
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nSaved later.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id, "persist": False}
    ).json()["run_id"]
    _drain_run_ws(client, run_id)

    resp = client.post(f"/runs/{run_id}/save")
    assert resp.status_code == 200, resp.text
    saved = resp.json()
    assert saved["kind"] == "result_md"
    assert saved["id"]

    # The document now appears in the list and the run carries output_doc_id.
    doc_ids = [d["id"] for d in client.get("/documents").json()]
    assert saved["id"] in doc_ids

    run = client.get(f"/runs/{run_id}").json()
    assert run["output_doc_id"] == saved["id"]

    out_doc = get_document(db, saved["id"])
    assert out_doc is not None
    assert out_doc.path == f"results/{saved['title']}.md"
    assert out_doc.path.startswith("results/")
    assert not out_doc.path.endswith(f"{saved['id']}.md")
    assert (Path(client.app.state.workspace) / out_doc.path).is_file()

    # Saving a second time is rejected (no duplicate document).
    resp2 = client.post(f"/runs/{run_id}/save")
    assert resp2.status_code == 409


def test_save_run_result_rewrites_obsidian_links(client, provider, db) -> None:
    linked_id = _upload(client, "ekonomiya-tokenov.md", b"token savings")
    linked = get_document(db, linked_id)
    assert linked is not None
    with db.connect() as conn:
        conn.execute(
            "UPDATE document SET title = ? WHERE id = ?",
            ("Экономия токенов", linked_id),
        )
    stem = Path(linked.path).stem

    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("См. [[Экономия токенов]].")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id, "persist": False}
    ).json()["run_id"]
    _drain_run_ws(client, run_id)

    run_before = client.get(f"/runs/{run_id}").json()
    assert run_before["result_text"] == "См. [[Экономия токенов]]."

    resp = client.post(f"/runs/{run_id}/save")
    assert resp.status_code == 200, resp.text
    saved = resp.json()

    run_after = client.get(f"/runs/{run_id}").json()
    assert run_after["result_text"] == "См. [[Экономия токенов]]."

    out_doc = get_document(db, saved["id"])
    assert out_doc is not None
    file_text = (Path(client.app.state.workspace) / out_doc.path).read_text(
        encoding="utf-8"
    )
    assert f"[[{stem}]]" in file_text
    assert "[[Экономия токенов]]" not in file_text
    input_doc = get_document(db, doc_id)
    assert input_doc is not None
    assert f"[[{Path(input_doc.path).stem}]]" in file_text


def test_save_run_result_ensures_parent_wikilinks(client, provider, db) -> None:
    doc_id = _upload(client, "cover-letter.md", b"source text")
    input_doc = get_document(db, doc_id)
    assert input_doc is not None
    input_stem = Path(input_doc.path).stem
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("Summary without wiki links.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id, "persist": False}
    ).json()["run_id"]
    _drain_run_ws(client, run_id)

    resp = client.post(f"/runs/{run_id}/save")
    assert resp.status_code == 200, resp.text
    saved = resp.json()

    out_doc = get_document(db, saved["id"])
    assert out_doc is not None
    file_text = (Path(client.app.state.workspace) / out_doc.path).read_text(
        encoding="utf-8"
    )
    assert "## Ссылки" in file_text
    assert f"- [[{input_stem}]]" in file_text
    run_after = client.get(f"/runs/{run_id}").json()
    assert run_after["result_text"] == "Summary without wiki links."


def test_save_run_result_attaches_to_session(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nSaved into session.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply",
        json={"doc_id": doc_id, "persist": False, "session_id": session_id},
    ).json()["run_id"]
    _drain_run_ws(client, run_id)

    saved = client.post(f"/runs/{run_id}/save").json()
    session_docs = list_session_documents(db, session_id)
    # The input doc is attached up front (session-scoped tools need it
    # visible), the saved result is attached after the fact.
    assert {d.id for d in session_docs} == {doc_id, saved["id"]}

    workspace = Path(client.app.state.workspace)
    tools = build_document_tools(db, workspace, session_id)

    async def _list():
        _, fn = tools.get("list_documents")
        return await fn()

    async def _read(doc_id_: str):
        _, fn = tools.get("read_document")
        return await fn(doc_id=doc_id_)

    listed_ids = {item["id"] for item in asyncio.run(_list())}
    assert listed_ids == {doc_id, saved["id"]}
    read = asyncio.run(_read(saved["id"]))
    assert "error" not in read
    assert "Saved into session" in read["text"]


def test_apply_persist_attaches_output_to_session_via_api(
    client, provider, db
) -> None:
    session_id = client.post("/sessions").json()["id"]
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nAttached via session_id.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply",
        json={"doc_id": doc_id, "persist": True, "session_id": session_id},
    ).json()["run_id"]
    finish = _drain_run_ws(client, run_id)[-1]
    assert finish["type"] == "finish"
    assert finish["status"] == "ok"
    assert finish["output_doc_id"] is not None

    from catalog.skills.repo_run import get_run

    assert get_run(db, run_id)["session_id"] == session_id
    session_docs = list_session_documents(db, session_id)
    # The input doc is attached up front (session-scoped tools need it
    # visible), the persisted output is attached after the fact.
    assert {d.id for d in session_docs} == {doc_id, finish["output_doc_id"]}

    workspace = Path(client.app.state.workspace)
    tools = build_document_tools(db, workspace, session_id)

    async def _list():
        _, fn = tools.get("list_documents")
        return await fn()

    async def _read(doc_id_: str):
        _, fn = tools.get("read_document")
        return await fn(doc_id=doc_id_)

    listed_ids = {item["id"] for item in asyncio.run(_list())}
    assert listed_ids == {doc_id, finish["output_doc_id"]}
    read = asyncio.run(_read(finish["output_doc_id"]))
    assert "error" not in read
    assert "Attached via session_id" in read["text"]


def test_apply_stream_tools_scoped_to_session(client, provider, db) -> None:
    session_id = client.post("/sessions").json()["id"]
    attached_id = _upload(client, "attached.md", b"visible in session")
    secret_id = _upload(client, "secret.md", b"not attached")
    attach_documents(db, session_id, [attached_id])

    skill_id = _seed_committed_skill(
        db,
        allowed_tools=["list_documents", "read_document"],
        verify_checks=[VerifyCheck("non_empty")],
        max_retries=2,
    )
    provider.script = [
        _completion(
            tool_calls=[
                ToolCall(id="list-1", name="list_documents", arguments={})
            ]
        ),
        _completion("# Result\n\nScoped listing ok."),
    ]

    run_id = client.post(
        f"/skills/{skill_id}/apply",
        json={
            "doc_ids": [attached_id],
            "persist": True,
            "session_id": session_id,
        },
    ).json()["run_id"]
    frames = _drain_run_ws(client, run_id)
    assert frames[-1]["type"] == "finish"
    assert frames[-1]["status"] == "ok"

    list_results = [
        f for f in frames if f.get("type") == "tool_result" and f.get("name") == "list_documents"
    ]
    assert list_results
    listed = json.loads(list_results[0]["result"])
    listed_ids = {item["id"] for item in listed}
    assert listed_ids == {attached_id}
    assert secret_id not in listed_ids


def test_apply_attaches_input_docs_to_session_before_run(client, provider, db) -> None:
    """A skills-panel input doc is not pre-attached to the planner session —
    apply must attach it itself so the session-scoped read_document tool
    (built from run["session_id"]) can actually see it."""
    session_id = client.post("/sessions").json()["id"]
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db,
        allowed_tools=["read_document"],
        verify_checks=[VerifyCheck("non_empty")],
        max_retries=0,
    )
    provider.script = [
        _completion(
            tool_calls=[
                ToolCall(id="read-1", name="read_document", arguments={"doc_id": doc_id})
            ]
        ),
        _completion("# Result\n\nRead the input document just fine."),
    ]

    run_id = client.post(
        f"/skills/{skill_id}/apply",
        json={"doc_ids": [doc_id], "persist": True, "session_id": session_id},
    ).json()["run_id"]
    frames = _drain_run_ws(client, run_id)

    read_results = [
        f for f in frames if f.get("type") == "tool_result" and f.get("name") == "read_document"
    ]
    assert read_results
    assert "document_not_available_in_session" not in read_results[0]["result"]

    finish = frames[-1]
    assert finish["type"] == "finish"
    assert finish["status"] == "ok"


def test_apply_missing_session_returns_404(client, db) -> None:
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=0
    )
    resp = client.post(
        f"/skills/{skill_id}/apply",
        json={"doc_ids": [doc_id], "session_id": "missing-session"},
    )
    assert resp.status_code == 404
    assert resp.json()["detail"] == "session not found"


def test_save_run_result_missing_run_returns_404(client, db) -> None:
    resp = client.post("/runs/does-not-exist/save")
    assert resp.status_code == 404


def test_save_run_result_already_persisted_returns_409(client, provider, db) -> None:
    """A persist=True run already has output_doc_id — saving again is rejected."""
    doc_id = _upload(client, "input.md", b"source text")
    skill_id = _seed_committed_skill(
        db, verify_checks=[VerifyCheck("non_empty")], max_retries=2
    )
    provider.script = [_completion("# Result\n\nAuto-saved.")]

    run_id = client.post(
        f"/skills/{skill_id}/apply", json={"doc_id": doc_id}
    ).json()["run_id"]
    _drain_run_ws(client, run_id)

    resp = client.post(f"/runs/{run_id}/save")
    assert resp.status_code == 409


def test_apply_arity_mismatch_returns_422(client, db) -> None:
    """A skill declaring input_arity=2 rejects a single-doc apply with 422."""
    doc_id = _upload(client, "input.md", b"source text")
    config = SkillConfig(
        name="Merger",
        description="needs exactly two docs",
        system_prompt="You merge two documents.",
        allowed_tools=["read_document"],
        model="test/model",
        max_iterations=4,
        max_retries=1,
        verify_checks=[VerifyCheck("non_empty")],
        input_arity=2,
    )
    skill_id = create_skill(
        db, name=config.name, description=config.description, config=config, status="committed"
    )

    resp = client.post(f"/skills/{skill_id}/apply", json={"doc_ids": [doc_id]})
    assert resp.status_code == 422
    assert "expects 2 input" in resp.json()["detail"]


def test_apply_arity_one_rejects_two_docs(client, db) -> None:
    doc_a = _upload(client, "a.md", b"alpha")
    doc_b = _upload(client, "b.md", b"beta")
    config = SkillConfig(
        name="Single",
        description="one doc only",
        system_prompt="Process one document.",
        allowed_tools=["read_document"],
        model="test/model",
        max_iterations=4,
        max_retries=1,
        verify_checks=[VerifyCheck("non_empty")],
        input_arity=1,
    )
    skill_id = create_skill(
        db, name=config.name, description=config.description, config=config, status="committed"
    )
    resp = client.post(
        f"/skills/{skill_id}/apply", json={"doc_ids": [doc_a, doc_b]}
    )
    assert resp.status_code == 422
    assert "expects 1 input" in resp.json()["detail"]


def test_apply_list_arity_accepts_multiple_docs(client, provider, db) -> None:
    doc_a = _upload(client, "a.md", b"alpha")
    doc_b = _upload(client, "b.md", b"beta")
    config = SkillConfig(
        name="Lister",
        description="any number of docs",
        system_prompt="Process the documents.",
        allowed_tools=["read_document"],
        model="test/model",
        max_iterations=4,
        max_retries=1,
        verify_checks=[VerifyCheck("non_empty")],
        input_arity=None,
    )
    skill_id = create_skill(
        db, name=config.name, description=config.description, config=config, status="committed"
    )
    provider.script = [_completion("# Result\n\nOk.")]
    resp = client.post(
        f"/skills/{skill_id}/apply", json={"doc_ids": [doc_a, doc_b]}
    )
    assert resp.status_code == 200, resp.text
    assert "run_id" in resp.json()


def test_apply_requires_at_least_one_doc(client, db) -> None:
    """An empty doc list (no doc_ids and no doc_id) is rejected with 422."""
    skill_id = _seed_committed_skill(db)
    resp = client.post(f"/skills/{skill_id}/apply", json={})
    assert resp.status_code == 422


def test_export_docx_200_writes_export_and_rescan_skips(client, db) -> None:
    from catalog.storage.repo_document import list_documents

    doc_id = _upload(client, "note.md", b"# Title\n\nHello\n")
    resp = client.post("/export/docx", json={"doc_ids": [doc_id], "title": "Report"})
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["ok"] is True
    assert body["path"].startswith("export/")
    assert body["path"].endswith(".docx")
    assert body["headings"] >= 1
    workspace = Path(client.app.state.workspace)
    dest = workspace / body["path"]
    assert dest.is_file()
    before = {d.path for d in list_documents(db)}
    assert body["path"] not in before
    again = client.post("/workspaces/rescan").json()
    assert body["path"] not in again["added"]
    after = {d.path for d in list_documents(db)}
    assert body["path"] not in after


def test_export_docx_404_unknown_document(client) -> None:
    resp = client.post("/export/docx", json={"doc_ids": ["missing-id"]})
    assert resp.status_code == 404
    assert resp.json()["detail"] == "document not found"


def test_export_docx_filter_allows_write_tool(client, db) -> None:
    workspace = Path(client.app.state.workspace)
    tools = build_document_tools(db, workspace)
    subset = tools.filter(["export_docx"])
    assert subset.names() == ["export_docx"]
    spec = subset.specs()[0]
    assert spec.side == "write"


def test_named_outputs_preview_save_creates_both_documents(client, db) -> None:
    workspace = Path(client.app.state.workspace)
    doc_id = _upload(client, "input.md", b"source text")
    skill = SkillConfig(
        name="splitter",
        description="two outputs",
        system_prompt="",
        allowed_tools=[],
        model="test/model",
        kind="script",
        code='result = {"brief": document.upper(), "table": "A -> a"}\n',
        outputs=[
            SkillOutput(key="brief", description="Текст"),
            SkillOutput(key="table", description="Таблица"),
        ],
        verify_checks=[VerifyCheck("non_empty")],
    )
    skill_id = create_skill(
        db,
        name=skill.name,
        description=skill.description,
        config=skill,
        status="committed",
    )
    result = asyncio.run(
        apply_skill_collect(
            provider=client.app.state.provider,
            db=db,
            workspace_dir=str(workspace),
            skill=skill,
            skill_id=skill_id,
            input_doc_ids=[doc_id],
            base_tools=build_document_tools(db, workspace),
            persist=False,
        )
    )
    assert result.status == "ok"
    assert result.output_doc_id is None
    run_id = result.run_id
    assert run_id is not None

    run = client.get(f"/runs/{run_id}").json()
    assert run["output_doc_id"] is None
    assert run["output_doc_ids"] == []
    assert run["result_artifacts"] == {"brief": "SOURCE TEXT", "table": "A -> a"}
    assert all(d["kind"] != "result_md" for d in client.get("/documents").json())

    resp = client.post(f"/runs/{run_id}/save")
    assert resp.status_code == 200, resp.text
    saved = resp.json()
    assert saved["kind"] == "result_md"
    run_after = client.get(f"/runs/{run_id}").json()
    assert run_after["output_doc_id"] == saved["id"]
    assert len(run_after["output_doc_ids"]) == 2
    assert run_after["output_doc_ids"][0] == saved["id"]
    docs = client.get("/documents").json()
    result_docs = [d for d in docs if d["kind"] == "result_md"]
    assert len(result_docs) == 2
    resp2 = client.post(f"/runs/{run_id}/save")
    assert resp2.status_code == 409
